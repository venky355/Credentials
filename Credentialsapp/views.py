from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth import authenticate, login, logout, update_session_auth_hash
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.shortcuts import render, redirect, get_object_or_404
from django.utils.html import mark_safe
from .forms import *
from .models import User, Product, Wishlist, Category
from .models import Product, Cart, CartItem
# from django.db.models import Sum, F, DecimalField
from django.urls import reverse
from django.http import JsonResponse
from .models import CartItem
from .models import UserProfile
from django.http.response import HttpResponse
# from django.core.exceptions import ObjectDoesNotExist
# import pandas as pd
# import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def main_home(request):
    return render(request, 'main_home.html')

def registration(request):
    if request.method == 'POST':
        form = RegistrationForm(request.POST)
        if form.is_valid():
            user = form.save(commit=False)
            password = form.cleaned_data['password']
            user.set_password(password)
            is_dealer = form.cleaned_data.get('is_dealer', False)
            if is_dealer:
                user.role = User.Role.DEALER
            else:
                user.role = User.Role.USERS
            
            user.save()
            login(request, user)
            if user.role == User.Role.DEALER:
                return redirect('product_list')  
            else:
                return redirect('user_home')  
    else:
        form = RegistrationForm()
    
    return render(request, 'registration.html', {'form': form})

def user_login(request):
    if request.method == 'POST':
        form = AuthenticationForm(request, data=request.POST)
        if form.is_valid():
            username = form.cleaned_data['username']
            password = form.cleaned_data['password']
            user = authenticate(request, username=username, password=password)
            if user is not None and user.is_active:
                login(request, user)
                if user.role == User.Role.DEALER:
                    return redirect('product_list')
                elif user.dealer_details == 'Category Management User':
                    return redirect('category_list')
                else:
                    return redirect('user_home')
            else:
                messages.error(request, 'Invalid username or password.')
    else:
        form = AuthenticationForm()
    return render(request, 'login.html', {'form': form})

@login_required
def category_list(request):
    categories = Category.objects.all()
    return render(request, 'category_list.html', {'categories': categories})

@login_required
def category_detail(request, category_id):
    category = get_object_or_404(Category, pk=category_id)
    products = Product.objects.filter(category=category)

    context = {
        'category': category,
        'products': products
    }
    return render(request, 'category_detail.html', context)
@login_required
def add_category(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        if name:
            Category.objects.create(name=name)
            return redirect('category_list')
    return render(request, 'add_category.html')

@login_required
def update_category(request, category_id):
    category = get_object_or_404(Category, pk=category_id)
    if request.method == 'POST':
        name = request.POST.get('name')
        if name:
            category.name = name
            category.save()
            return redirect('category_list')
    return render(request, 'update_category.html', {'category': category})

@login_required
def delete_category(request, category_id):
    category = get_object_or_404(Category, pk=category_id)
    if request.method == 'POST':
        category.delete()
        return redirect('category_list')
    return render(request, 'delete_category.html', {'category': category})

@login_required
def user_home(request):
    if request.user.is_authenticated:
        user_wishlist_ids = Wishlist.objects.filter(user=request.user).values_list('product__id', flat=True)
        products = Product.objects.all()
        categories = Category.objects.all()

        context = {
            'products': products,
            'categories': categories,
            'user_wishlist_ids': user_wishlist_ids,
        }

        return render(request, 'user_home.html', context)
    else:
        return render(request, 'login.html')
    

@login_required
def change_password(request):
    if request.method == 'POST':
        form = CustomPasswordChangeForm(user=request.user, data=request.POST)
        if form.is_valid():
            user = form.save()
            update_session_auth_hash(request, user)
            return redirect('home')
    else:
        form = CustomPasswordChangeForm(user=request.user)
    return render(request, 'change_password.html', {'form': form})

@login_required
def update_user_details(request):
    user = request.user
    if request.method == 'POST':
        form = UserProfileForm(request.POST, instance=user)
        if form.is_valid():
            form.save()
            messages.success(request, 'Your details were successfully updated!')
            return redirect('user_home')
    else:
        form = UserProfileForm(instance=user)
    return render(request, 'update_user_details.html', {'form': form})

@login_required
def product_list(request):
    products = Product.objects.filter(dealer=request.user)
    return render(request, 'product_list.html', {'products': products})

@login_required
def add_product(request):
    if request.method == 'POST':
        form = ProductForm(request.POST, request.FILES)
        if form.is_valid():
            product = form.save(commit=False)
            product.dealer = request.user
            product.is_approved = False  
            product.save()
            messages.success(request, 'Product added successfully. Waiting for admin approval.')
            return redirect('product_list')
        else:
            messages.error(request, 'Form is invalid. Please correct the errors.')
    else:
        form = ProductForm()
    return render(request, 'add_product.html', {'form': form})

def upload_excel(request):
    if request.method == 'POST':
        excel_file = request.FILES.get('excel_file')

        if not excel_file.name.endswith(('.xlsx', '.xls')):
            return render(request, 'upload_excel.html', {'error': 'File is not in Excel format'})

        try:
            wb = load_workbook(excel_file)
            if 'Sheet1' in wb.sheetnames:
                worksheet = wb['Sheet1']
            else:
                return render(request, 'upload_excel.html', {'error': 'Worksheet "Sheet1" not found in the Excel file'})

            excel_data = []
            for row in worksheet.iter_rows(values_only=True):
                excel_data.append(row)

            return render(request, 'upload_excel.html', {'excel_data': excel_data})

        except Exception as e:
            return render(request, 'upload_excel.html', {'error': f'Error processing Excel file: {str(e)}'})

    return render(request, 'upload_excel.html')


# def import_from_excel(request):
#     if request.method == 'POST' and request.FILES['excel_file']:
#         excel_file = request.FILES['excel_file']

#         try:
#             # Read the Excel file into a DataFrame
#             df = pd.read_excel(excel_file)

#             # Iterate over each row in the DataFrame
#             for index, row in df.iterrows():
#                 product_name = row['Product Name']
#                 category_name = row['Category Name']
#                 country = row['Country']
#                 quantity = row['Quantity']
#                 price = row['Price']
#                 image_url = row['Image']

#                 # Get or create Category based on category_name
#                 category, created = Category.objects.get_or_create(name=category_name)

#                 # Save or update the Product based on product_name
#                 product, created = Product.objects.update_or_create(
#                     name=product_name,
#                     defaults={
#                         'category': category,
#                         'country': country,
#                         'quantity': quantity,
#                         'price': price,
#                         'image': image_url if image_url else None
#                     }
#                 )

#             return HttpResponse('Import successful!')

#         except Exception as e:
#             return HttpResponse(f'Import failed. Error: {str(e)}')

#     return render(request, 'import_from_excel.html')

# def export_to_excel(request):
#     # Query all products from the database
#     products = Product.objects.all()

#     # Create a DataFrame with product data
#     product_data = {
#         'Product Name': [product.name for product in products],
#         'Category Name': [product.category.name for product in products],
#         'Country': [product.country for product in products],
#         'Quantity': [product.quantity for product in products],
#         'Price': [product.price for product in products],
#         'Image': [product.image.url if product.image else None for product in products]
#     }
#     df = pd.DataFrame(product_data)

#     # Create response object to return the Excel file as a download
#     response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
#     response['Content-Disposition'] = 'attachment; filename=product_list.xlsx'

#     # Write DataFrame to Excel file and attach to response
#     df.to_excel(response, index=False)

#     return response

@login_required
def update_product(request, product_id):
    product = get_object_or_404(Product, id=product_id)

    if request.method == 'POST':
        form = ProductForm(request.POST, instance=product)
        if form.is_valid():
            form.save()
            messages.success(request, 'Product updated successfully.')
            return redirect('product_list')
        else:
            messages.error(request, 'Error updating product. Please correct the form.')

    else:
        form = ProductForm(instance=product)

    return render(request, 'update_product.html', {'form': form, 'product': product})

@login_required
def delete_product(request, product_id):
    product = get_object_or_404(Product, pk=product_id, dealer=request.user)
    if request.method == 'POST':
        product.delete()
        return redirect('product_list')
    return render(request, 'delete_product.html', {'product': product})

@login_required
def product_search(request):
    query = request.GET.get('query')

    if query:
        # Search for products by name or category name containing the query
        products = Product.objects.filter(name__icontains=query) | Product.objects.filter(category__name__icontains=query)
        
        highlighted_products = []
        for product in products:
            product_name = product.name
            highlighted_text = product_name.replace(query, f'<span style="background-color: yellow;">{query}</span>')
            highlighted_products.append({
                'product': product,
                'highlighted_name': mark_safe(highlighted_text),
                'highlighted_category': None,  # Placeholder for highlighted category
                'related_products': product.category.product_set.exclude(id=product.id)[:3] if product.category else []
            })
    else:
        # If no query is provided, return all products
        products = Product.objects.all()
        highlighted_products = [{
            'product': product,
            'highlighted_name': product.name,
            'highlighted_category': None,
            'related_products': []
        } for product in products]

    context = {
        'highlighted_products': highlighted_products,
        'query': query
    }
    return render(request, 'search_results.html', context)

@login_required
def wishlist(request):
    wishlist_items = Wishlist.objects.filter(user=request.user)
    return render(request, 'wishlist.html', {'wishlist_items': wishlist_items})

@login_required
def add_to_wishlist(request, product_id):
    product = get_object_or_404(Product, pk=product_id)
    existing_wishlist_item = Wishlist.objects.filter(user=request.user, product=product).first()

    if existing_wishlist_item:
        messages.info(request, 'This item is already in your wishlist.')
        return redirect('wishlist')

    if request.method == 'POST':
        form = WishlistForm(request.POST)
        if form.is_valid():
            wishlist_item = form.save(commit=False)
            wishlist_item.user = request.user
            wishlist_item.product = product
            wishlist_item.save()
            messages.success(request, 'Item added to wishlist successfully.')
            return redirect('wishlist')
    else:
        form = WishlistForm()
    return render(request, 'add_to_wishlist.html', {'form': form, 'product': product})

@login_required
def remove_from_wishlist(request, wishlist_item_id):
    wishlist_item = get_object_or_404(Wishlist, id=wishlist_item_id)
    if request.method == 'POST':
        wishlist_item.delete()
        return redirect('wishlist')


@login_required
def logout_view(request):
    logout(request)
    return redirect('main_home')


@login_required
def view_cart(request):
    try:
        cart = Cart.objects.get(user=request.user)
        cart_items = CartItem.objects.filter(cart=cart)
        total_cost = sum(item.item_total_cost() for item in cart_items)
        return render(request, 'view_cart.html', {'cart_items': cart_items, 'total_cost': total_cost})
    except Cart.DoesNotExist:
        messages.warning(request, "Cart does not exist. Please add items to your cart.")
        return redirect('user_home')
    
@login_required
def add_to_cart(request, product_id):
    product = get_object_or_404(Product, pk=product_id)
    if not hasattr(request.user, 'cart'):
        Cart.objects.create(user=request.user)
    cart = request.user.cart
    cart_item, created = CartItem.objects.get_or_create(cart=cart, product=product)
    if Wishlist.objects.filter(user=request.user, product=product).exists():
        wishlist_item = Wishlist.objects.get(user=request.user, product=product)
        wishlist_item.delete()
    if not created:
        cart_item.quantity += 1
        cart_item.save()

    return redirect('view_cart')

@login_required
def remove_from_cart(request, cart_item_id):
    cart_item = get_object_or_404(CartItem, pk=cart_item_id)
    cart_item.delete()
    return redirect('view_cart')

@login_required
def checkout(request):
    user = request.user
    try:
        user_profile = UserProfile.objects.get(user=user)
    except UserProfile.DoesNotExist:
        user_profile = UserProfile.objects.create(user=user)

    cart = Cart.objects.get(user=user)
    total_cost = sum(item.item_total_cost() for item in cart.cart_items.all())

    if request.method == 'POST':
        return redirect(reverse('payment_success'))

    context = {
        'cart': cart,
        'total_cost': total_cost,
        'user_address': user_profile.address if user_profile else None
    }
    return render(request, 'checkout.html', context)

def update_cart_item(request, cart_item_id):
    if request.method == 'POST' and request.is_ajax():
        quantity = request.POST.get('quantity')
        cart_item = get_object_or_404(CartItem, id=cart_item_id)
        cart_item.quantity = quantity
        cart_item.save()
        item_total_cost = cart_item.product.price * cart_item.quantity
        total_cost = cart_item.cart.total_cost()
        response_data = {
            'item_total_cost': item_total_cost,
            'total_cost': total_cost
        }

        return JsonResponse(response_data)
    return JsonResponse({'error': 'Invalid request'}, status=400)

@login_required
def view_profile(request):
    user_profile = UserProfile.objects.get_or_create(user=request.user)[0]
    return render(request, 'view_profile.html', {'user_profile': user_profile})

@login_required
def edit_profile(request):
    user_profile = UserProfile.objects.get_or_create(user=request.user)[0]
    if request.method == 'POST':
        form = ProfileForm(request.POST, instance=user_profile)
        if form.is_valid():
            form.save()
            return redirect('checkout')
    else:
        form = ProfileForm(instance=user_profile)
    return render(request, 'edit_profile.html', {'form': form})

@login_required
def update_address(request):
    if request.method == 'POST':
        address = request.POST.get('address')
        user_profile = UserProfile.objects.get_or_create(user=request.user)[0]
        user_profile.address = address
        user_profile.save()
        return redirect('view_profile')  
    else:
        return redirect('view_profile') 
    

@login_required
# @admin_required
def pending_products(request):
    pending_products = Product.objects.filter(is_approved=False)
    return render(request, 'pending_products.html', {'pending_products': pending_products})


@login_required
# @admin_required
def approve_product(request, product_id):
    product = get_object_or_404(Product, pk=product_id)
    product.is_approved = True
    product.save()
    return HttpResponse('Product approved')

@login_required
def reject_product(request, product_id):
    product = get_object_or_404(Product, pk=product_id)
    product.is_approved = False
    product.save()
    return HttpResponse('Product Rejected')

@login_required    
def toggle_wishlist(request):
    if request.method == 'POST' and request.is_ajax():
        product_id = request.POST.get('product_id')
        in_wishlist = request.POST.get('in_wishlist')

        product = get_object_or_404(Product, id=product_id)
        user = request.user  # Assuming you have a user associated with the request

        if in_wishlist == 'true':
            # Remove product from wishlist
            Wishlist.objects.filter(user=user, product=product).delete()
            return JsonResponse({'in_wishlist': False})
        else:
            # Add product to wishlist
            Wishlist.objects.create(user=user, product=product)
            return JsonResponse({'in_wishlist': True})

    return JsonResponse({'error': 'Invalid request'}, status=400)  

@login_required
def update_address(request):
    if request.method == 'POST' and request.is_ajax():
        new_address = request.POST.get('new_address')

        # Update the user's address
        user_profile = UserProfile.objects.get_or_create(user=request.user)[0]
        user_profile.address = new_address
        user_profile.save()

        return JsonResponse({'success': True})
    else:
        return JsonResponse({'success': False, 'message': 'Invalid request'})
    
def process_payment(request):
    if request.method == 'POST':
        return JsonResponse({'success': True})
    else:
        return JsonResponse({'success': False})
    
def payment_success(request):
    return render(request, 'payment_success.html')  
  
def process_payment(request):
    return redirect('payment_success')

def wishlist_actions(request):
    if request.method == 'POST':
        action = request.POST.get('action')
        product_id = request.POST.get('product_id')
        if action == 'add':
            return JsonResponse({'status': 'added'})

        elif action == 'remove':
            return JsonResponse({'status': 'removed'})

    return JsonResponse({'status': 'error'}, status=400)